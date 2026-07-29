from __future__ import annotations

import io
import math
import re
import statistics
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from typing import Any

import msoffcrypto
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook

app = FastAPI()

EXPORT_COLUMNS = [
    "member_name",
    "email",
    "Care_Email",
    "contact_number",
    "Alternate_Contact",
    "Transaction_Date",
    "transaction_amount",
    "transaction_id",
    "passing_year",
    "course",
]
POLICY_COLUMN = "Policy (New/Renewal)"

ALIASES = {
    "member_name": ["member name", "member_name", "full name", "customer name", "student name", "name"],
    "email": ["email", "email id", "email address", "personal email"],
    "Care_Email": ["care email", "care_email", "care email id", "care email address"],
    "contact_number": ["contact number", "contact_number", "contact no", "mobile number", "mobile no", "phone number"],
    "Alternate_Contact": ["alternate contact", "alternate_contact", "alternate number", "alternate mobile", "secondary contact"],
    "Transaction_Date": ["transaction date", "transaction_date", "payment date", "txn date", "date of transaction"],
    "transaction_amount": ["transaction amount", "transaction_amount", "payment amount", "paid amount", "txn amount", "amount"],
    "transaction_id": ["transaction id", "transaction_id", "txn id", "payment id", "reference id", "utr", "receipt number"],
    POLICY_COLUMN: ["policy new renewal", "policy (new renewal)", "new renewal", "new/renewal", "policy status", "new or renewal"],
    "passing_year": ["passing year", "passing_year", "batch", "batch year", "year of passing", "graduation year", "passout year"],
    "course": ["course", "course name", "program", "programme", "program name"],
}

EXTRA_ALIASES = {
    "dob": ["dob", "date of birth", "birth date"],
    "age": ["age", "member age"],
    "gender": ["gender", "sex"],
    "city": ["city", "town", "member city"],
    "state": ["state", "province", "member state"],
    "country": ["country", "nation"],
    "pincode": ["pincode", "pin code", "postal code", "zip code", "member pincode"],
    "sum_insured": ["sum insured", "sum_insured", "sum assured", "coverage amount", "cover amount"],
    "premium": ["premium inc gst", "premium_inc_gst", "premium including gst", "gross premium", "premium"],
    "insurer": ["insurer", "insurance company", "carrier"],
    "plan_name": ["plan name", "plan_name", "insurance plan", "product name"],
    "policy_type": ["policy type", "policy_type", "insurance type"],
    "policy_name": ["policy name", "policy_name"],
    "pay_mode": ["pay mode", "pay_mode", "payment mode"],
    "relationship": ["relationship", "member relationship"],
    "nominee_relationship": [
        "nominee relationship",
        "nominee_relationship",
        "nominee relation",
        "relation with nominee",
        "relationship with nominee",
    ],
}


def normalize(value: Any) -> str:
    text = "" if value is None else str(value).strip().lower().replace("&", " and ")
    text = re.sub(r"[()\[\]{}\\/\-]+", " ", text)
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "null", "nat"}:
        return ""
    return re.sub(r"\s+", " ", text)


def clean_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = clean_text(value).replace(",", "")
    negative = text.startswith("(") and text.endswith(")")
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in {"", "-", ".", "-."}:
        return None
    try:
        result = float(text)
        return -result if negative else result
    except ValueError:
        return None


def clean_phone(value: Any) -> str:
    digits = re.sub(r"\D", "", clean_text(value))
    if not digits or set(digits) == {"0"}:
        return ""
    if digits.startswith("0091") and len(digits) >= 14:
        digits = digits[4:]
    elif digits.startswith("91") and len(digits) >= 12:
        digits = digits[2:]
    elif digits.startswith("0") and len(digits) == 11:
        digits = digits[1:]
    return digits[-10:] if len(digits) > 10 else digits


def parse_date(value: Any) -> date | None:
    if value in (None, "", 0, "0", "0.0"):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and 1 <= float(value) <= 100000:
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    text = clean_text(value)
    for pattern in (
        "%d/%b/%y", "%d/%b/%Y", "%d-%b-%y", "%d-%b-%Y",
        "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y",
        "%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%m-%d-%Y",
    ):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def normalize_policy(value: Any) -> str:
    normalized = normalize(value)
    if normalized in {"new", "n", "new_policy", "new_member"}:
        return "New"
    if normalized in {"renewal", "renew", "renewed", "r", "policy_renewal"}:
        return "Renewal"
    return ""


def decrypt_workbook(raw: bytes, password: str) -> bytes:
    if raw[:2] == b"PK":
        return raw
    try:
        encrypted = msoffcrypto.OfficeFile(io.BytesIO(raw))
        encrypted.load_key(password=password)
        output = io.BytesIO()
        encrypted.decrypt(output)
        return output.getvalue()
    except Exception as exc:
        raise HTTPException(400, f"Could not unlock workbook. Check password. ({exc})") from exc


def age_band(age: int) -> str:
    if age < 18:
        return "Under 18"
    if age <= 25:
        return "18–25"
    if age <= 35:
        return "26–35"
    if age <= 45:
        return "36–45"
    if age <= 55:
        return "46–55"
    if age <= 65:
        return "56–65"
    return "66+"


def premium_band(value: float) -> str:
    if value < 10000:
        return "Below ₹10,000"
    if value < 25000:
        return "₹10,000–₹24,999"
    if value < 50000:
        return "₹25,000–₹49,999"
    if value < 100000:
        return "₹50,000–₹99,999"
    return "₹100,000+"


def grouped(counter: Counter, totals: defaultdict | None = None, *, sort_by_count: bool = True) -> list[dict[str, Any]]:
    pairs = counter.most_common() if sort_by_count else list(counter.items())
    output: list[dict[str, Any]] = []
    for label, count in pairs:
        row: dict[str, Any] = {"label": str(label), "count": int(count)}
        if totals is not None:
            amount = round(float(totals[label]), 2)
            row.update(amount=amount, average=round(amount / count, 2) if count else 0)
        output.append(row)
    return output


def natural_batch_key(label: str) -> tuple[int, str]:
    match = re.search(r"\d{4}", label)
    return (int(match.group()) if match else 999999, label.lower())


@app.post("/")
async def analyze(files: list[UploadFile] = File(...), password: str = Form("")):
    raw_rows: list[dict[str, Any]] = []
    processing_log: list[dict[str, Any]] = []

    normalized_aliases = {
        key: {normalize(alias) for alias in values + [key]}
        for key, values in ALIASES.items()
    }
    normalized_extras = {
        key: {normalize(alias) for alias in values + [key]}
        for key, values in EXTRA_ALIASES.items()
    }

    for upload in files:
        raw = await upload.read()
        if len(raw) > 15 * 1024 * 1024:
            raise HTTPException(413, f"{upload.filename} exceeds 15 MB")
        try:
            workbook = load_workbook(
                io.BytesIO(decrypt_workbook(raw, password)),
                read_only=True,
                data_only=True,
            )
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(400, f"Unable to read {upload.filename}: {exc}") from exc

        sheets_used = 0
        rows_extracted = 0
        for sheet in workbook.worksheets:
            best: tuple[int, int, list[Any], list[str]] | None = None
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=1, max_row=min(20, sheet.max_row), values_only=True),
                start=1,
            ):
                headers = [normalize(value) for value in row]
                score = sum(
                    any(header in alias_set for alias_set in normalized_aliases.values())
                    or any(header in alias_set for alias_set in normalized_extras.values())
                    for header in headers
                )
                if any(header in normalized_aliases["Transaction_Date"] for header in headers):
                    score += 5
                if best is None or score > best[0]:
                    best = (score, row_number, list(row), headers)

            if not best or best[0] < 5:
                continue

            _, header_row, _, normalized_headers = best
            required_mapping: dict[str, int] = {}
            extra_mapping: dict[str, int] = {}
            for key, alias_set in normalized_aliases.items():
                for index, header in enumerate(normalized_headers):
                    if header in alias_set:
                        required_mapping[key] = index
                        break
            for key, alias_set in normalized_extras.items():
                for index, header in enumerate(normalized_headers):
                    if header in alias_set:
                        extra_mapping[key] = index
                        break

            if "Transaction_Date" not in required_mapping or len(required_mapping) < 4:
                continue

            sheets_used += 1
            for source_row_number, row in enumerate(
                sheet.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                if not any(value not in (None, "") for value in row):
                    continue

                def get(index: int | None):
                    return row[index] if index is not None and index < len(row) else None

                record = {key: get(required_mapping.get(key)) for key in EXPORT_COLUMNS + [POLICY_COLUMN]}
                record.update({key: get(index) for key, index in extra_mapping.items()})
                record["_source_file"] = upload.filename or "workbook"
                record["_source_sheet"] = sheet.title
                record["_source_row"] = source_row_number
                raw_rows.append(record)
                rows_extracted += 1

        processing_log.append(
            {
                "file": upload.filename,
                "sheets_used": sheets_used,
                "rows_extracted": rows_extracted,
            }
        )

    if not raw_rows:
        raise HTTPException(400, "No usable data sheet found")

    rows: list[dict[str, Any]] = []
    invalid_dates = 0
    exact_duplicates = 0
    duplicate_ids = 0
    exact_seen: set[tuple[Any, ...]] = set()
    transaction_ids: set[str] = set()
    policy_present = False

    for source in raw_rows:
        transaction_date = parse_date(source.get("Transaction_Date"))
        if not transaction_date:
            invalid_dates += 1
            continue

        transaction_id = clean_text(source.get("transaction_id"))
        if transaction_id.endswith(".0") and transaction_id[:-2].isdigit():
            transaction_id = transaction_id[:-2]

        policy = normalize_policy(source.get(POLICY_COLUMN))
        policy_present = policy_present or bool(policy)
        export_row = {
            "member_name": clean_text(source.get("member_name")),
            "email": clean_text(source.get("email")).lower(),
            "Care_Email": clean_text(source.get("Care_Email")).lower(),
            "contact_number": clean_phone(source.get("contact_number")),
            "Alternate_Contact": clean_phone(source.get("Alternate_Contact")),
            "Transaction_Date": transaction_date.isoformat(),
            "transaction_amount": round(clean_number(source.get("transaction_amount")) or 0, 2),
            "transaction_id": transaction_id,
            POLICY_COLUMN: policy,
            "passing_year": clean_text(source.get("passing_year")),
            "course": clean_text(source.get("course")),
        }

        exact_key = tuple(export_row[column] for column in EXPORT_COLUMNS + [POLICY_COLUMN])
        if exact_key in exact_seen:
            exact_duplicates += 1
            continue
        exact_seen.add(exact_key)

        if transaction_id and transaction_id in transaction_ids:
            duplicate_ids += 1
            continue
        if transaction_id:
            transaction_ids.add(transaction_id)

        premium = clean_number(source.get("premium"))
        if premium in (None, 0):
            premium = export_row["transaction_amount"]
        sum_insured = clean_number(source.get("sum_insured"))

        age = None
        dob = parse_date(source.get("dob"))
        age_value = clean_number(source.get("age"))
        if dob:
            age = transaction_date.year - dob.year - (
                (transaction_date.month, transaction_date.day) < (dob.month, dob.day)
            )
        elif age_value is not None and 0 <= age_value <= 120:
            age = int(age_value)

        product_parts: list[str] = []
        for key in ("plan_name", "policy_type", "policy_name"):
            value = clean_text(source.get(key))
            if value and value not in product_parts:
                product_parts.append(value)

        rows.append(
            {
                "export": export_row,
                "date": transaction_date,
                "amount": export_row["transaction_amount"],
                "premium": round(float(premium or 0), 2),
                "sum_insured": sum_insured,
                "age": age,
                "state": clean_text(source.get("state")).title(),
                "city": clean_text(source.get("city")).title(),
                "country": clean_text(source.get("country")).upper(),
                "pincode": clean_text(source.get("pincode")),
                "course": export_row["course"],
                "passing_year": export_row["passing_year"],
                "insurance_product": " • ".join(product_parts),
                "insurer": clean_text(source.get("insurer")),
                "gender": clean_text(source.get("gender")).title(),
                "relationship": clean_text(source.get("relationship")).title(),
                "nominee_relationship": clean_text(source.get("nominee_relationship")).title(),
                "pay_mode": clean_text(source.get("pay_mode")).title(),
                "policy": policy,
                "source_file": source.get("_source_file", ""),
                "source_sheet": source.get("_source_sheet", ""),
                "source_row": source.get("_source_row", ""),
            }
        )

    if not rows:
        raise HTTPException(400, "No rows remain after Transaction Date cleaning")

    export_columns = EXPORT_COLUMNS.copy()
    if policy_present:
        export_columns.insert(8, POLICY_COLUMN)

    cleaned_rows = [
        {column: item["export"].get(column, "") for column in export_columns}
        for item in rows
    ]

    daily = defaultdict(lambda: [0, 0.0, 0.0])
    monthly = defaultdict(lambda: [0, 0.0, 0.0])
    policy_counter = Counter()
    sum_counter = Counter()
    sum_amount = defaultdict(float)
    sum_premium = defaultdict(float)
    premium_counter = Counter()
    premium_total = defaultdict(float)
    age_counter = Counter()
    age_amount = defaultdict(float)
    state_counter = Counter()
    state_amount = defaultdict(float)
    city_counter = Counter()
    city_amount = defaultdict(float)
    pincode_counter = Counter()
    pincode_amount = defaultdict(float)
    course_counter = Counter()
    course_amount = defaultdict(float)
    batch_counter = Counter()
    batch_amount = defaultdict(float)
    product_counter = Counter()
    product_amount = defaultdict(float)
    insurer_counter = Counter()
    insurer_amount = defaultdict(float)
    gender_counter = Counter()
    relationship_counter = Counter()
    nominee_counter = Counter()
    pay_mode_counter = Counter()

    for item in rows:
        day_key = item["date"].isoformat()
        month_key = item["date"].strftime("%Y-%m")
        for bucket, key in ((daily, day_key), (monthly, month_key)):
            bucket[key][0] += 1
            bucket[key][1] += item["amount"]
            bucket[key][2] += item["premium"]

        if item["policy"]:
            policy_counter[item["policy"]] += 1
        if item["sum_insured"] and item["sum_insured"] > 0:
            label = f"₹{item['sum_insured']:,.0f}"
            sum_counter[label] += 1
            sum_amount[label] += item["amount"]
            sum_premium[label] += item["premium"]
        if item["premium"] > 0:
            label = premium_band(item["premium"])
            premium_counter[label] += 1
            premium_total[label] += item["premium"]
        if item["age"] is not None:
            label = age_band(item["age"])
            age_counter[label] += 1
            age_amount[label] += item["amount"]

        for value, counter, totals in (
            (item["state"], state_counter, state_amount),
            (item["city"], city_counter, city_amount),
            (item["pincode"], pincode_counter, pincode_amount),
            (item["course"], course_counter, course_amount),
            (item["passing_year"], batch_counter, batch_amount),
            (item["insurance_product"], product_counter, product_amount),
            (item["insurer"], insurer_counter, insurer_amount),
        ):
            if value:
                counter[value] += 1
                totals[value] += item["amount"]

        if item["gender"]:
            gender_counter[item["gender"]] += 1
        if item["relationship"]:
            relationship_counter[item["relationship"]] += 1
        if item["nominee_relationship"]:
            nominee_counter[item["nominee_relationship"]] += 1
        if item["pay_mode"]:
            pay_mode_counter[item["pay_mode"]] += 1

    daily_rows = [
        {"label": key, "period": key, "count": value[0], "amount": round(value[1], 2), "premium": round(value[2], 2)}
        for key, value in sorted(daily.items())
    ]
    monthly_rows = [
        {"label": key, "period": key, "count": value[0], "amount": round(value[1], 2), "premium": round(value[2], 2)}
        for key, value in sorted(monthly.items())
    ]
    sum_rows = [
        {
            "label": label,
            "count": count,
            "amount": round(sum_amount[label], 2),
            "premium": round(sum_premium[label], 2),
            "average": round(sum_premium[label] / count, 2) if count else 0,
        }
        for label, count in sum_counter.most_common()
    ]
    premium_order = [
        "Below ₹10,000",
        "₹10,000–₹24,999",
        "₹25,000–₹49,999",
        "₹50,000–₹99,999",
        "₹100,000+",
    ]
    premium_rows = [
        {
            "label": label,
            "count": premium_counter[label],
            "amount": round(premium_total[label], 2),
            "premium": round(premium_total[label], 2),
            "average": round(premium_total[label] / premium_counter[label], 2),
        }
        for label in premium_order
        if premium_counter[label]
    ]
    age_order = ["Under 18", "18–25", "26–35", "36–45", "46–55", "56–65", "66+"]
    age_rows = [
        {"label": label, "count": age_counter[label], "amount": round(age_amount[label], 2)}
        for label in age_order
        if age_counter[label]
    ]

    batch_rows = grouped(batch_counter, batch_amount, sort_by_count=False)
    batch_rows.sort(key=lambda item: natural_batch_key(item["label"]))

    total_amount = round(sum(item["amount"] for item in rows), 2)
    total_premium = round(sum(item["premium"] for item in rows), 2)
    top_sum = sum_rows[0] if sum_rows else None
    top_state = grouped(state_counter, state_amount)[:1]
    top_city = grouped(city_counter, city_amount)[:1]
    top_pincode = grouped(pincode_counter, pincode_amount)[:1]
    top_course = grouped(course_counter, course_amount)[:1]
    top_product = grouped(product_counter, product_amount)[:1]
    top_batch = max(batch_rows, key=lambda item: item["amount"], default=None)
    top_nominee = grouped(nominee_counter)[:1]

    insights: list[str] = []
    if top_sum:
        insights.append(f"{top_sum['label']} is the most selected sum insured with {top_sum['count']} selections.")
    if premium_rows:
        most_frequent = max(premium_rows, key=lambda item: item["count"])
        highest_value = max(premium_rows, key=lambda item: item["premium"])
        insights.append(f"{most_frequent['label']} is the most frequently purchased premium band.")
        insights.append(f"{highest_value['label']} contributes the highest premium value.")
    if top_state:
        insights.append(f"{top_state[0]['label']} is the leading state with {top_state[0]['count']} transactions.")
    if top_city:
        insights.append(f"{top_city[0]['label']} is the leading city with {top_city[0]['count']} transactions.")
    if top_pincode:
        insights.append(f"Pincode {top_pincode[0]['label']} has the highest transaction count ({top_pincode[0]['count']}).")
    if top_course:
        insights.append(f"{top_course[0]['label']} is the leading course segment.")
    if top_batch:
        insights.append(
            f"Batch {top_batch['label']} contributes the highest transaction amount at ₹{top_batch['amount']:,.2f}."
        )
    if top_nominee:
        insights.append(f"{top_nominee[0]['label']} is the most common nominee relationship.")
    if monthly_rows:
        best_month = max(monthly_rows, key=lambda item: item["amount"])
        insights.append(f"{best_month['label']} is the highest-value transaction month at ₹{best_month['amount']:,.2f}.")
    if daily_rows:
        best_day = max(daily_rows, key=lambda item: item["amount"])
        insights.append(f"{best_day['label']} is the highest-value transaction date at ₹{best_day['amount']:,.2f}.")
    if policy_counter:
        total_valid_policy = sum(policy_counter.values())
        insights.append(
            f"Renewals represent {policy_counter.get('Renewal', 0) / total_valid_policy * 100:.1f}% of rows with a valid New/Renewal value."
        )

    return {
        "meta": {
            "export_columns": export_columns,
            "policy_included": policy_present,
            "processed_at": datetime.now().isoformat(timespec="seconds"),
            "files_processed": len(files),
            "row_order": "source_order",
        },
        "kpis": {
            "total_records": len(rows),
            "unique_members": len({item["export"]["member_name"].lower() for item in rows if item["export"]["member_name"]}),
            "total_transaction_amount": total_amount,
            "average_transaction_amount": round(total_amount / len(rows), 2),
            "median_transaction_amount": round(statistics.median(item["amount"] for item in rows), 2),
            "total_premium": total_premium,
            "average_premium": round(total_premium / len(rows), 2),
            "new_count": policy_counter.get("New", 0),
            "renewal_count": policy_counter.get("Renewal", 0),
            "most_selected_sum_insured": top_sum["label"] if top_sum else "Not available",
            "top_state": top_state[0]["label"] if top_state else "Not available",
            "top_city": top_city[0]["label"] if top_city else "Not available",
            "top_course": top_course[0]["label"] if top_course else "Not available",
            "top_insurance_product": top_product[0]["label"] if top_product else "Not available",
        },
        "cleaned_rows": cleaned_rows,
        "analysis": {
            "daily_trend": daily_rows,
            "monthly_trend": monthly_rows,
            "policy": grouped(policy_counter),
            "sum_insured": sum_rows,
            "premium_bands": premium_rows,
            "age": age_rows,
            "state": grouped(state_counter, state_amount),
            "city": grouped(city_counter, city_amount),
            "pincode": grouped(pincode_counter, pincode_amount),
            "course": grouped(course_counter, course_amount),
            "passing_year": batch_rows,
            "insurance_products": grouped(product_counter, product_amount),
            "insurers": grouped(insurer_counter, insurer_amount),
            "gender": grouped(gender_counter),
            "relationship": grouped(relationship_counter),
            "nominee_relationship": grouped(nominee_counter),
            "pay_mode": grouped(pay_mode_counter),
        },
        "insights": insights,
        "data_quality": {
            "rows_before_cleaning": len(raw_rows),
            "invalid_dates_removed": invalid_dates,
            "exact_duplicates_removed": exact_duplicates,
            "duplicate_transaction_ids_removed": duplicate_ids,
            "final_rows": len(rows),
            "processing_log": processing_log,
            "export_order": "Original source order retained; no date or amount sorting applied.",
        },
    }
