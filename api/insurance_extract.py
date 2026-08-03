from __future__ import annotations

import csv
import io
import re
from typing import Any

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook

from api.insurance_config import ALIASES, EXTRA, EXPORT_COLUMNS, POLICY_COLUMN
from api.insurance_utils import decrypt, norm, number, parse_date, phone, policy_value, text


def _report_year(value: Any) -> int | None:
    match = re.search(r"(?:19|20)\d{2}", text(value))
    if not match:
        return None
    year = int(match.group(0))
    return year if 1900 <= year <= 2100 else None


def _open_sheets(raw: bytes, filename: str, password: str):
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        decoded = None
        for encoding in ("utf-8-sig", "utf-8", "cp1252"):
            try:
                decoded = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if decoded is None:
            raise HTTPException(400, f"Unable to decode {filename} as CSV")
        sample = decoded[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        return [list(csv.reader(io.StringIO(decoded), dialect))]

    if lower_name.endswith(".xls"):
        try:
            import xlrd
        except ImportError as exc:
            raise HTTPException(500, "Legacy .xls support is not installed") from exc
        try:
            workbook = xlrd.open_workbook(file_contents=raw)
        except Exception:
            workbook = xlrd.open_workbook(file_contents=decrypt(raw, password))
        return [
            [sheet.row_values(row_number) for row_number in range(sheet.nrows)]
            for sheet in workbook.sheets()
        ]

    workbook = load_workbook(
        io.BytesIO(decrypt(raw, password)),
        read_only=True,
        data_only=True,
    )
    return workbook.worksheets


async def extract_rows(
    files: list[UploadFile],
    password: str,
    labels: list[str],
    college: str,
    file_years: list[str] | None = None,
):
    alias = {key: {norm(item) for item in values + [key]} for key, values in ALIASES.items()}
    extra = {key: {norm(item) for item in values + [key]} for key, values in EXTRA.items()}
    raw_rows, logs = [], []
    source_sequence = 0

    for file_index, upload in enumerate(files):
        raw = await upload.read()
        if len(raw) > 15 * 1024 * 1024:
            raise HTTPException(413, f"{upload.filename} exceeds 15 MB")

        fallback_plan = text(labels[file_index]) if file_index < len(labels) else ""
        fallback_plan = fallback_plan or f"Data Set {file_index + 1}"
        year_override = _report_year(
            file_years[file_index]
            if file_years and file_index < len(file_years)
            else ""
        )

        try:
            worksheets = _open_sheets(raw, upload.filename or "report.xlsx", password)
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(400, f"Unable to read {upload.filename}: {exc}") from exc

        used = extracted = 0
        detected_plans: set[str] = set()

        for worksheet in worksheets:
            best = None
            is_materialized = isinstance(worksheet, list)
            preview_rows = (
                worksheet[:15]
                if is_materialized
                else worksheet.iter_rows(
                    min_row=1,
                    max_row=min(15, worksheet.max_row or 15),
                    values_only=True,
                )
            )
            for row_number, row in enumerate(preview_rows, 1):
                headers = [norm(item) for item in row]
                score = sum(
                    any(header in values for values in alias.values())
                    or any(header in values for values in extra.values())
                    for header in headers
                )
                score += 5 * any(header in alias["Transaction_Date"] for header in headers)
                score += 4 * any(header in alias["transaction_amount"] for header in headers)
                if best is None or score > best[0]:
                    best = (score, row_number, headers)

            if not best:
                continue

            _, header_row, headers = best
            mapping, dimensions = {}, {}

            for key, values in alias.items():
                for index, header in enumerate(headers):
                    if header in values:
                        mapping[key] = index
                        break

            for key, values in extra.items():
                for index, header in enumerate(headers):
                    if header in values:
                        dimensions[key] = index
                        break

            # A year-labelled report can still be compared when it has no row-level
            # date column. In that case 1 January of the selected report year is used
            # only as an analysis anchor and is reported in data-quality controls.
            has_date = "Transaction_Date" in mapping
            has_amount = "transaction_amount" in mapping
            minimum_score = 4 if year_override else 5
            if best[0] < minimum_score or not has_amount or (not has_date and not year_override):
                continue

            used += 1
            data_rows = (
                worksheet[header_row:]
                if is_materialized
                else worksheet.iter_rows(min_row=header_row + 1, values_only=True)
            )
            for row in data_rows:
                if not any(value not in (None, "") for value in row):
                    continue

                get = lambda idx: row[idx] if idx is not None and idx < len(row) else None
                record = {key: get(mapping.get(key)) for key in EXPORT_COLUMNS + [POLICY_COLUMN]}
                record.update({key: get(index) for key, index in dimensions.items()})

                detected_plan = text(record.get("plan_name"))
                if detected_plan:
                    detected_plans.add(detected_plan)

                record.update(
                    _fallback_plan=fallback_plan,
                    _college=text(college),
                    _file=upload.filename or f"File {file_index + 1}",
                    _year_override=year_override,
                    _policy_column_present=POLICY_COLUMN in mapping,
                    _sequence=source_sequence,
                )
                raw_rows.append(record)
                source_sequence += 1
                extracted += 1

        logs.append(
            {
                "file": upload.filename,
                "fallback_label": fallback_plan,
                "detected_plans": sorted(detected_plans),
                "report_year": year_override,
                "sheets_used": used,
                "rows_extracted": extracted,
            }
        )

    return raw_rows, logs


def clean_rows(raw_rows: list[dict[str, Any]], strict_single: bool = False):
    return _clean_ordered_rows(raw_rows, comparison=not strict_single)


def _duplicate_value(column: str, value: Any) -> str:
    """Return a stable value for duplicate checks without source metadata."""
    if column in {"Transaction_Date", "dob"}:
        parsed = parse_date(value)
        if parsed:
            return parsed.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return text(value).casefold()


def _raw_row_key(raw: dict[str, Any]) -> tuple[str, ...]:
    source_columns = EXPORT_COLUMNS + [POLICY_COLUMN, *EXTRA.keys()]
    return tuple(
        _duplicate_value(column, raw.get(column)) for column in source_columns
    )


def _identifier_key(value: Any) -> str:
    return text(value).casefold()


def _normalised_transaction_id(value: Any) -> str:
    transaction_id = text(value)
    if transaction_id.endswith(".0") and transaction_id[:-2].isdigit():
        transaction_id = transaction_id[:-2]
    return transaction_id


def _cleaning_scope(raw: dict[str, Any], comparison: bool) -> tuple[str, str]:
    """Keep comparison-year reports independent while combining split files."""
    if not comparison:
        return ("single", "single")
    year = str(_report_year(raw.get("_year_override")) or "")
    label = norm(text(raw.get("_fallback_plan")) or "unspecified_plan")
    return (label, year or text(raw.get("_file")))


def _clean_ordered_rows(raw_rows: list[dict[str, Any]], comparison: bool):
    """Apply the same ordered business cleaning workflow in both modes."""
    ordered_raw_rows = sorted(raw_rows, key=lambda item: item.get("_sequence", 0))
    has_policy = any(bool(raw.get("_policy_column_present")) for raw in raw_rows)

    # Rule 1: remove blank/zero/invalid dates, then duplicate rows.
    exact_seen = set()
    stage_one: list[tuple[dict[str, Any], Any]] = []
    exact_duplicates = 0
    invalid_dates = 0
    for raw in ordered_raw_rows:
        transaction_date = parse_date(raw.get("Transaction_Date"))
        if not transaction_date:
            invalid_dates += 1
            continue
        scope = _cleaning_scope(raw, comparison)
        exact_key = (scope, *_raw_row_key(raw))
        if exact_key in exact_seen:
            exact_duplicates += 1
            continue
        exact_seen.add(exact_key)
        stage_one.append((raw, transaction_date, scope))

    # Rule 2: on the remaining data, remove repeated Transaction IDs globally.
    transaction_ids = set()
    stage_two: list[tuple[dict[str, Any], Any, str]] = []
    duplicate_transaction_ids = 0
    for raw, transaction_date, scope in stage_one:
        transaction_id = _normalised_transaction_id(raw.get("transaction_id"))
        transaction_id_key = (scope, transaction_id.casefold())
        if transaction_id and transaction_id_key in transaction_ids:
            duplicate_transaction_ids += 1
            continue
        if transaction_id:
            transaction_ids.add(transaction_id_key)
        stage_two.append((raw, transaction_date, transaction_id, scope))

    # Rule 3: after date/row/transaction-ID cleaning, each nonblank email ID
    # may occur only once in an uploaded report.  Email and Care Email are
    # checked independently; requiring Member + Email + Care Email to match
    # allowed repeated email IDs to inflate enrolment and premium totals.
    email_seen = set()
    care_email_seen = set()
    stage_three = []
    duplicate_emails = 0
    duplicate_care_emails = 0
    for raw, transaction_date, transaction_id, scope in stage_two:
        email = _identifier_key(raw.get("email"))
        care_email = _identifier_key(raw.get("Care_Email"))
        email_key = (scope, email)
        care_email_key = (scope, care_email)

        if email and email_key in email_seen:
            duplicate_emails += 1
            continue
        if care_email and care_email_key in care_email_seen:
            duplicate_care_emails += 1
            continue

        if email:
            email_seen.add(email_key)
        if care_email:
            care_email_seen.add(care_email_key)
        stage_three.append((raw, transaction_date, transaction_id))

    rows = []
    canonical_plans: dict[str, str] = {}
    for raw, transaction_date, transaction_id in stage_three:
        transaction_amount = number(raw.get("transaction_amount"))
        override_year = _report_year(raw.get("_year_override"))
        analysis_year = str(override_year or transaction_date.year)
        policy = policy_value(raw.get(POLICY_COLUMN))
        exported = {
            "member_name": text(raw.get("member_name")),
            "email": text(raw.get("email")).lower(),
            "Care_Email": text(raw.get("Care_Email")).lower(),
            # Rule 4: phone country codes are removed only after all duplicate checks.
            "contact_number": phone(raw.get("contact_number")),
            "Alternate_Contact": phone(raw.get("Alternate_Contact")),
            "Transaction_Date": transaction_date.isoformat(),
            "transaction_amount": round(transaction_amount or 0, 2),
            "transaction_id": transaction_id,
            POLICY_COLUMN: policy,
            "passing_year": text(raw.get("passing_year")),
            "course": text(raw.get("course")),
        }

        detected_plan = text(raw.get("plan_name")) or text(raw.get("policy_name"))
        fallback_plan = text(raw.get("_fallback_plan"))
        fallback_key = norm(fallback_plan)
        explicit_common_plan = fallback_plan and fallback_key not in {
            "unspecified_plan",
        } and not fallback_key.startswith("data_set_")
        supplied_plan = (
            fallback_plan
            if explicit_common_plan
            else detected_plan or fallback_plan or "Unspecified Plan"
        )
        plan_key = norm(supplied_plan)
        plan = canonical_plans.setdefault(plan_key, supplied_plan)

        sum_insured = number(raw.get("sum_insured"))
        age = None
        dob = parse_date(raw.get("dob"))
        given_age = number(raw.get("age"))
        if dob:
            age = transaction_date.year - dob.year - (
                (transaction_date.month, transaction_date.day) < (dob.month, dob.day)
            )
        elif given_age is not None and 0 <= given_age <= 120:
            age = int(given_age)

        products = []
        for key in ["plan_name", "policy_type", "policy_name"]:
            value = text(raw.get(key))
            if value and value not in products:
                products.append(value)

        rows.append(
            {
                "export": exported,
                "sequence": raw.get("_sequence", 0),
                "file": raw.get("_file", ""),
                "plan": plan,
                "college": raw.get("_college", ""),
                "date": transaction_date,
                "year": analysis_year,
                "month_number": transaction_date.month,
                "month_label": transaction_date.strftime("%b"),
                "month_period": f"{transaction_date.strftime('%b')} {analysis_year}",
                "amount": exported["transaction_amount"],
                "premium": exported["transaction_amount"],
                "sum_insured": sum_insured,
                "age": age,
                "gender": text(raw.get("gender")).title(),
                "state": text(raw.get("state")).title(),
                "city": text(raw.get("city")).title(),
                "country": text(raw.get("country")).title(),
                "pincode": text(raw.get("pincode")),
                "course": exported["course"],
                "passing_year": exported["passing_year"],
                "insurance_product": " • ".join(products),
                "insurer": text(raw.get("insurer")),
                "payment_mode": text(raw.get("pay_mode")).title(),
                "nominee_relationship": text(raw.get("nominee_relationship")).title(),
                "policy": policy,
            }
        )

    return rows, has_policy, {
        "invalid_dates_removed": invalid_dates,
        "exact_duplicates_removed": exact_duplicates,
        "duplicate_transaction_ids_removed": duplicate_transaction_ids,
        "duplicate_member_names_removed": 0,
        "duplicate_emails_removed": duplicate_emails,
        "duplicate_care_emails_removed": duplicate_care_emails,
        "duplicate_identity_rows_removed": duplicate_emails + duplicate_care_emails,
        "report_year_overrides_applied": sum(
            1 for raw, *_ in stage_three if _report_year(raw.get("_year_override"))
        ),
        "dates_inferred_from_report_year": 0,
    }
